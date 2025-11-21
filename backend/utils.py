import re
from openpyxl.utils import get_column_letter

from styles import DISABLED_FILL, GRAY, GRAYB


def norm_token(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^\w\-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def tree_max_depth(nodes, level=0):
    if not nodes:
        return level - 1
    m = level
    for n in nodes:
        m = max(m, tree_max_depth(n.get("children") or [], level + 1))
    return m


def autosize_columns(ws, min_w=10, max_w=120):
    for col in ws.columns:
        col = list(col)
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))


def gray_out_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        bold = bool(cell.font and cell.font.bold)
        cell.fill = DISABLED_FILL
        cell.font = GRAYB if bold else GRAY


def gray_out_row_full(ws, row, total_cols):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        bold = bool(cell.font and cell.font.bold)
        cell.fill = DISABLED_FILL
        cell.font = GRAYB if bold else GRAY


def strip_prefix_levels(nodes, n):
    cur = nodes
    for _ in range(n):
        if not cur:
            return []
        first = cur[0]
        cur = first.get("children") or []
    return cur
