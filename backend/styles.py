from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

ORANGE = PatternFill("solid", fgColor="EA5B2B")
CYAN   = PatternFill("solid", fgColor="63D3FF")
PALEGR = PatternFill("solid", fgColor="D8F4D2")
BLUE   = PatternFill("solid", fgColor="2F78BD")
LIME   = PatternFill("solid", fgColor="CCFF66")
TITLE_GRAY = PatternFill("solid", fgColor="D1D5DB")
DISABLED_FILL = PatternFill("solid", fgColor="E5E7EB")
YELLOW = PatternFill("solid", fgColor="FFF2CC")

WHITEB = Font(color="FFFFFF", bold=True)
BLACK  = Font(color="000000")
BLACKB = Font(color="000000", bold=True)
GRAY   = Font(color="808080")
GRAYB  = Font(color="808080", bold=True)

LEFT   = Alignment(horizontal="left",  vertical="center", wrap_text=False)
CENTER = Alignment(horizontal="center",vertical="center", wrap_text=False)

GRID_GRAY    = Side(style="thin", color="B0B0B0")
THICK        = Side(style="thick", color="000000")
THICK_BOTTOM = Side(style="thick", color="000000")
THICK_TOP    = Side(style="thick", color="000000")

BOX = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
BOX2 = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
