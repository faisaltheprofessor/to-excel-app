from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse
import io, re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI()

# ===== Config =====
PREFIX = "203_"
GROUPS_FROM_LEVEL = 3
SHEET_NAME  = "GE_Gruppenstruktur"
SHEET2_NAME = "Strukt. Ablage Behörde"
SHEET3_NAME = "Geschäftsrollen"

ROW_BAND, ROW_TITLE, ROW_CAPTION, ROW_HEADERS = 2, 3, 3, 5
DATA_START_ROW = 4

# The PHP side wraps real data with 3 ancestors (.PANKOW → ba → DigitaleAkte-203).
# For Sheets 2 & 3 we skip those wrappers.
SKIP_PARENTS = 3

ORG_NAME = "BA-PANKOW"
ADMIN_ACCOUNT = "T1_PL_extMA_DigitaleAkte_Fach_Admin_Role"

# ===== Styles =====
ORANGE = PatternFill("solid", fgColor="EA5B2B")
CYAN   = PatternFill("solid", fgColor="63D3FF")
PALEGR = PatternFill("solid", fgColor="D8F4D2")   # highlight rows such as Ltg/Allg & Ablagen
BLUE   = PatternFill("solid", fgColor="2F78BD")   # containers/headers
LIME   = PatternFill("solid", fgColor="CCFF66")   # leaves
TITLE_GRAY = PatternFill("solid", fgColor="D1D5DB")
DISABLED_FILL = PatternFill("solid", fgColor="E5E7EB")
YELLOW = PatternFill("solid", fgColor="FFF2CC")

WHITEB = Font(color="FFFFFF", bold=True)
BLACK  = Font(color="000000")
BLACKB = Font(color="000000", bold=True)
GRAY   = Font(color="808080")
GRAYB  = Font(color="808080", bold=True)

LEFT   = Alignment(horizontal="left",  vertical="center", wrap_text=False)
CENTER = Alignment(horizontal="center",vertical="center", wrap_text=False)

GRID_GRAY    = Side(style="thin", color="B0B0B0")
THIN_BLACK   = Side(style="thin", color="000000")
THICK        = Side(style="thick", color="000000")
THICK_BOTTOM = Side(style="thick", color="000000")
THICK_TOP    = Side(style="thick", color="000000")

BOX = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
BOX2 = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

PERM_HEADERS = [
    "Lesen", "Schreiben", "Administrieren", "Löschadministration", "Ablageadministration",
    "Aktenplanadministration", "Vorlagenadministration", "Aussonderung",
    "Postverteilung- zentral", "Postverteilung- dezentral", "Designkonfiguration",
]
PERM_SUFFIX = ["RO", "", "FA", "LA", "AA", "APA", "VA", "AUS", "POZ", "POD", "DK"]

LEFT_HEADERS = [h for h in PERM_HEADERS if h != "Schreiben"]
LEFT_SUFFIX  = [s for h, s in zip(PERM_HEADERS, PERM_SUFFIX) if h != "Schreiben"]

FORCE_BLUE = {"Org", "Org Name"}

# Roles sheet config (overridden by rolesCount from request if provided)
ROLES_COUNT = 10
ROLE_HEADER_TEXT = "Rollenname"

# ---------------------------------------------------------------------------
# Utility helpers
# ---------------------------------------------------------------------------

def norm_token(s: str) -> str:
    """Normalize to a token (letters/digits/underscore/hyphen) for group keys."""
    s = (s or "").strip()
    s = re.sub(r"[^\w\-]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s

def tree_max_depth(nodes, level=0):
    """Return maximum depth, with root level = 0."""
    if not nodes:
        return level - 1
    m = level
    for n in nodes:
        m = max(m, tree_max_depth(n.get("children") or [], level + 1))
    return m

# ---------------------------------------------------------------------------
# Sheet 1: GE_Gruppenstruktur (with ASCII tree)
# ---------------------------------------------------------------------------

def create_sheet(last_name_col:int, perm1_cols:int, perm2_cols:int, max_depth:int):
    """
    Sheet layout:
      [tree columns][name] [spacer] [perm block 1] [spacer] [perm block 2]
      [spacer] [flat label] [spacer] [tree drawing columns]
    """
    spacer1 = last_name_col + 1
    perm1_s = spacer1 + 1
    perm1_e = perm1_s + perm1_cols - 1
    spacer2 = perm1_e + 1
    perm2_s = spacer2 + 1
    perm2_e = perm2_s + perm2_cols - 1
    spacer3 = perm2_e + 1
    flat_col = spacer3 + 1
    spacer4 = flat_col + 1
    tree_base = spacer4 + 1
    tree_end  = tree_base + max_depth

    wb = Workbook(); ws = wb.active; ws.title = SHEET_NAME

    for c in range(1, last_name_col):
        ws.column_dimensions[get_column_letter(c)].width = 4
    ws.column_dimensions[get_column_letter(last_name_col)].width = 26
    for c in [spacer1, spacer2, spacer3, spacer4]:
        ws.column_dimensions[get_column_letter(c)].width = 2
    for c in list(range(perm1_s, perm1_e+1)) + list(range(perm2_s, perm2_e+1)):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions[get_column_letter(flat_col)].width = 26
    for c in range(tree_base, tree_end+1):
        ws.column_dimensions[get_column_letter(c)].width = 6

    for r,h in [(ROW_BAND,24),(ROW_TITLE,20),(ROW_HEADERS-1,18)]:
        ws.row_dimensions[r].height = h

    ws.merge_cells(start_row=ROW_BAND, start_column=perm1_s, end_row=ROW_BAND, end_column=perm1_e)
    c1 = ws.cell(row=ROW_BAND, column=perm1_s, value="eDir")
    c1.fill = ORANGE; c1.font = WHITEB; c1.alignment = CENTER; c1.border = BOX

    chip = ws.cell(row=ROW_BAND, column=perm2_s, value="eDir")
    chip.fill = ORANGE; chip.font = WHITEB; chip.alignment = CENTER; chip.border = BOX

    ws.merge_cells(start_row=ROW_BAND, start_column=perm2_s+1, end_row=ROW_BAND, end_column=perm2_e)
    c2 = ws.cell(row=ROW_BAND, column=perm2_s+1, value="AD(DFSW)")
    c2.fill = CYAN; c2.font = BLACKB; c2.alignment = CENTER; c2.border = BOX

    ws.cell(row=ROW_TITLE, column=1, value="Struktur Gruppen mit Schreibzugriff").font = BLACKB
    ws.merge_cells(start_row=ROW_CAPTION, start_column=perm1_s, end_row=ROW_CAPTION, end_column=perm1_e)
    ws.cell(row=ROW_CAPTION, column=perm1_s, value="auf Knoten zusätzlich anzulegende Gruppen").font = BLACKB

    for j,hdr in enumerate(LEFT_HEADERS, start=perm1_s):
        hc = ws.cell(row=ROW_HEADERS, column=j, value=hdr)
        hc.font = BLACKB; hc.alignment = CENTER; hc.border = BOX
    for j,hdr in enumerate(PERM_HEADERS, start=perm2_s):
        hc = ws.cell(row=ROW_HEADERS, column=j, value=hdr)
        hc.font = BLACKB; hc.alignment = CENTER; hc.border = BOX

    ws.merge_cells(start_row=ROW_BAND, start_column=flat_col, end_row=ROW_BAND, end_column=tree_end)
    ab = ws.cell(row=ROW_BAND, column=flat_col, value="nscale strukturierte Ablage")
    ab.fill = PALEGR; ab.font = BLACKB; ab.alignment = LEFT; ab.border = BOX

    ws.cell(row=ROW_HEADERS-1, column=flat_col, value="Liste").font = BLACKB
    ws.merge_cells(start_row=ROW_HEADERS-1, start_column=tree_base, end_row=ROW_HEADERS-1, end_column=tree_end)
    bh = ws.cell(row=ROW_HEADERS-1, column=tree_base, value="Baum")
    bh.font = BLACKB; bh.alignment = LEFT

    return wb, ws, {
        "spacer1": spacer1, "perm1_s": perm1_s, "perm1_e": perm1_e,
        "spacer2": spacer2, "perm2_s": perm2_s, "perm2_e": perm2_e,
        "spacer3": spacer3, "flat_col": flat_col, "spacer4": spacer4,
        "tree_base": tree_base, "tree_end": tree_end, "max_depth": max_depth
    }

# ─────────────────────────────────────────────────────────────────────────────
# ASCII tree drawer — **key idea: shifted ancestor logic**
#
# For a row at depth L, columns BEFORE the elbow (0..L-2) should continue '│'
# if the **next ancestor** (index d+1) is NOT last among its siblings.
# Example: under IKT → SG1 subtree, column 0 must show '│' as long as SG1
# isn’t the last top-level child — independent of whether IKT has siblings.
# ─────────────────────────────────────────────────────────────────────────────

def draw_verticals_before_elbow(ws, row: int, level: int, last_stack, base_col: int = 1):
    """
    Draw '│' for columns 0..level-2.
    Decision at column d uses last_stack[d+1] (shifted by one).
    """
    if level <= 1:
        return
    for d in range(0, level - 1):                  # 0..L-2
        idx = d + 1                                # look at the *next* ancestor
        if idx < len(last_stack) and not last_stack[idx]:
            c = ws.cell(row=row, column=base_col + d, value="│")
            c.alignment = CENTER; c.font = GRAY

def draw_verticals_through_subtree(ws, row: int, level: int, last_stack, base_col: int = 1):
    """
    Continue verticals through a node's subtree region.
    We draw columns 0..level-1 (includes the parent's elbow column).
    Decision at column d uses last_stack[d+1].
    """
    if level <= 0:
        return
    for d in range(0, level):                       # 0..L-1
        idx = d + 1
        if idx < len(last_stack) and not last_stack[idx]:
            c = ws.cell(row=row, column=base_col + d, value="│")
            c.alignment = CENTER; c.font = GRAY

def draw_connectors(ws, row: int, level: int, last_stack, base_col: int = 1):
    """
    Draw ancestor verticals before the elbow + the elbow itself.
    Elbow is └ if current node is last; else ├.
    """
    if level <= 0:
        return
    draw_verticals_before_elbow(ws, row, level, last_stack, base_col=base_col)

    elbow_col = base_col + (level - 1)
    elbow = "└" if last_stack[-1] else "├"
    c = ws.cell(row=row, column=elbow_col, value=elbow)
    c.alignment = CENTER; c.font = GRAY

def extend_connectors(ws, start_row: int, end_row: int, level: int, last_stack, base_col: int = 1):
    """
    After drawing a node, extend verticals down across its subtree rows.
    We include the elbow column here (columns 0..level-1).
    """
    if end_row < start_row or level < 0:
        return
    for r in range(start_row, end_row + 1):
        draw_verticals_through_subtree(ws, r, level, last_stack, base_col=base_col)

# ─────────────────────────────────────────────────────────────────────────────

def style_cell_like_node(cell, label:str, is_container:bool):
    cell.alignment = LEFT; cell.border = BOX
    if is_container or label in FORCE_BLUE:
        cell.fill = BLUE; cell.font = WHITEB
    else:
        cell.fill = LIME; cell.font = BLACKB

def style_name_cell(cell, node_name:str, has_children:bool):
    cell.alignment = LEFT; cell.border = BOX
    if has_children or node_name in FORCE_BLUE:
        cell.fill = BLUE; cell.font = WHITEB
    else:
        cell.fill = LIME; cell.font = BLACKB

def is_ablg(s: str) -> bool:
    return (s or "").strip().lower() == "ablgoe"

def is_poeing(s: str) -> bool:
    return (s or "").strip().lower() == "poeing"

def gray_out_row(ws, row, col_start, col_end):
    """Gray out a range, preserving bold where present."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        bold = bool(cell.font and cell.font.bold)
        cell.fill = DISABLED_FILL
        cell.font = GRAYB if bold else GRAY

def write_rows(ws, nodes, row, level, last_stack, cols, lineage_names, lineage_apps):
    """
    Preorder traversal that writes each node, draws connectors,
    and fills group columns. **No spacer rows** are inserted.
    """
    name_col = level + 1
    for i, node in enumerate(nodes):
        name = (node.get("name") or "").strip()
        appn = (node.get("appName") or name).strip()
        children = node.get("children") or []
        if not name and not children:
            continue

        is_last = (i == len(nodes) - 1)
        disabled = (node.get("enabled") is False)

        # Draw connectors for this row
        draw_connectors(ws, row, level, last_stack + [is_last], base_col=1)

        # Name cell
        nc = ws.cell(row=row, column=name_col, value=name if name else "(unnamed)")
        style_name_cell(nc, name, bool(children))

        # Fill dashes between name and spacer
        for dc in range(name_col + 1, cols["spacer1"]):
            d = ws.cell(row=row, column=dc, value="-")
            d.alignment = CENTER; d.border = BOX

        # Clear spacers
        for sc in [cols["spacer1"], cols["spacer2"], cols["spacer3"], cols["spacer4"]]:
            ws.cell(row=row, column=sc, value="")

        # Left permission block (from certain depth)
        if level >= GROUPS_FROM_LEVEL and name:
            for j, suf in enumerate(LEFT_SUFFIX, start=cols["perm1_s"]):
                val = f"{name}-{suf}" if suf else name
                g = ws.cell(row=row, column=j, value=val)
                g.alignment = LEFT; g.border = BOX

        # Right permission block (group keys)
        if level >= GROUPS_FROM_LEVEL and name:
            tokens = lineage_names + [norm_token(name)]
            start = min(GROUPS_FROM_LEVEL, len(tokens) - 1)
            key = PREFIX + "_".join(tokens[start:])
            for j, suf in enumerate(PERM_SUFFIX, start=cols["perm2_s"]):
                val = f"{key}-{suf}" if suf else key
                g = ws.cell(row=row, column=j, value=val)
                g.alignment = LEFT; g.border = BOX

        # Flat label + tree columns on the right
        flat_label = appn if appn else (name if name else "(unnamed)")
        fc = ws.cell(row=row, column=cols["flat_col"], value=flat_label)
        style_cell_like_node(fc, flat_label, bool(children))

        r_name_col = cols["tree_base"] + level
        tv_label = appn if appn else (name if name else "(unnamed)")
        tv = ws.cell(row=row, column=r_name_col, value=tv_label)
        style_cell_like_node(tv, tv_label, bool(children))

        if disabled:
            gray_out_row(ws, row, 1, cols["tree_end"])

        # Mark subtree start
        subtree_start_row = row
        row += 1

        # Recurse
        if children:
            row = write_rows(
                ws, children, row, level + 1, last_stack + [is_last], cols,
                lineage_names + [norm_token(name)],
                lineage_apps + [appn]
            )

        # Extend verticals across the entire subtree (no gaps)
        subtree_end_row = row - 1
        extend_connectors(ws, subtree_start_row + 1, subtree_end_row, level, last_stack + [is_last], base_col=1)

    return row

def autosize_columns(ws, min_w=10, max_w=120):
    """Simple autosize with bounds."""
    for col in ws.columns:
        col = list(col)
        letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))

# ---------------------------------------------------------------------------
# Sheet 2: Strukt. Ablage Behörde
# ---------------------------------------------------------------------------

def strip_prefix_levels(nodes, n):
    cur = nodes
    for _ in range(n):
        if not cur:
            return []
        first = cur[0]
        cur = first.get("children") or []
    return cur

def allowed_types_for(label, typ):
    lab = (label or "").lower()
    if typ == "Posteingang" or lab.startswith("pe_") or "poeing" in lab:
        return "GOV_WORKING_FOLDER_INBOX"
    return "GOV_FILE" if typ == "Aktenablage" else ""

def break_inheritance_from_node(node, desc_text):
    v = node.get("unterbrechen")
    if isinstance(v, str) and v.strip().lower() in {"wahr", "true", "ja"}:
        return "WAHR"
    if v is True:
        return "WAHR"
    if "ohne leitungszugriff" in (desc_text or "").lower():
        return "WAHR"
    return ""

def build_group_key_from_names(path_names):
    tokens = [norm_token(x) for x in path_names if (x or "").strip()]
    return PREFIX + "_".join(tokens) if tokens else ""

def set_row_top_thick(ws, row, col_start=1, col_end=11):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(
            left=b.left or Side(style="thin", color="000000"),
            right=b.right or Side(style="thin", color="000000"),
            top=THICK_TOP,
            bottom=b.bottom or Side(style="thin", color="000000"),
        )

def set_row_bottom_thick(ws, row, col_start=1, col_end=11):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(
            left=b.left or Side(style="thin", color="000000"),
            right=b.right or Side(style="thin", color="000000"),
            top=b.top or Side(style="thin", color="000000"),
            bottom=THICK_BOTTOM,
        )

def set_group_bottom_thick(ws, row_last, col_start=1, col_end=11):
    set_row_bottom_thick(ws, row_last, col_start, col_end)

def make_addr(group_key: str, suffix: str = "") -> str:
    key = f"{group_key}-{suffix}" if suffix else group_key
    return f"{key}@{ORG_NAME}"

def is_poe_label(label: str) -> bool:
    low = (label or "").lower()
    return low.startswith("pe_") or "poeing" in low

def write_group_recursive(ws, node, r, path_names, path_apps, poeings):
    """
    Flatten to 'Strukt. Ablage Behörde'. Capture each PoEing plus its disabled state
    so we can mirror that gray state onto the derived PoKorb rows later.
    """
    name = (node.get("name") or "").strip()
    appn = (node.get("appName") or name).strip()
    kids = node.get("children") or []
    desc = (node.get("description") or "").strip()

    label = appn if appn else (name if name else "(unnamed)")
    is_parent = bool(kids)
    parent_raw = path_apps[-1] if path_apps else ""

    display_a = label
    parent_display = parent_raw
    typ = "Hierarchieelement" if is_parent else "Aktenablage"

    full_names_path = path_names + [name]
    perm_key = build_group_key_from_names(full_names_path)
    ancestor_keys = [build_group_key_from_names(full_names_path[:i]) for i in range(1, len(full_names_path)+1)]

    values = [display_a, desc, parent_display, typ,
              break_inheritance_from_node(node, desc),
              allowed_types_for(label, "Posteingang" if is_poe_label(label) else typ)]
    for j, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = LEFT
        c.border = BOX2
    ws.cell(row=r, column=1).fill = YELLOW

    g_list = [make_addr(perm_key, "RO")]
    h_list = [make_addr(perm_key, "")]
    i_list = [make_addr(perm_key, "FA")]
    j_list = [make_addr(k, "LA") for k in ancestor_keys]
    k_list = [make_addr(perm_key, "AA")]

    cells_lists = [g_list, h_list, i_list, j_list, k_list]
    for offset, items in enumerate(cells_lists, start=7):
        cell_val = ";".join(items) + f";{ADMIN_ACCOUNT}"
        c = ws.cell(row=r, column=offset, value=cell_val)
        c.alignment = LEFT
        c.border = BOX2

    ws.cell(row=r, column=1).font = (BLACKB if is_parent else BLACK)
    for j in range(2, 12):
        ws.cell(row=r, column=j).font = BLACK
    if is_parent:
        set_row_top_thick(ws, r, 1, 11)

    if node.get("enabled") is False:
        gray_out_row(ws, r, 1, 11)

    current_row = r

    # Capture PoEing nodes and whether they were disabled
    if is_poe_label(label):
        poeings.append({"path": list(path_apps), "disabled": (node.get("enabled") is False)})

    for child in kids:
        current_row += 1
        current_row, _ = write_group_recursive(
            ws,
            child,
            current_row,
            path_names = full_names_path,
            path_apps  = path_apps + [label],
            poeings    = poeings
        )

    if is_parent:
        set_group_bottom_thick(ws, current_row, 1, 11)

    return current_row, poeings

def add_second_sheet(wb: Workbook, tree):
    ws = wb.create_sheet(title=SHEET2_NAME)

    headers = [
        "Bezeichnung strukturierte Ablage",
        "Beschreibung",
        "Eltern strukturierte Ablage",
        "Art",
        "Vererbung aus übergeordnetem Element unterbrechen",
        "Erlaubte Aktentypen",
        "Lesen", "Schreiben", "Administrieren", "Löschadministration", "Ablageadministration"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = BLACKB; c.alignment = LEFT; c.border = BOX2

    for col_idx in range(7, 12):
        ws.cell(row=1, column=col_idx).fill = PALEGR

    working_nodes = strip_prefix_levels(tree, SKIP_PARENTS)

    r = 2
    all_poeings = []

    for top in working_nodes:
        r, poe = write_group_recursive(ws, top, r, path_names=[], path_apps=[], poeings=[])
        all_poeings.extend(poe)
        r += 1

    # Append PoKorb rows (mirror disabled state from the corresponding PoEing).
    for pe in all_poeings:
        parent_app = pe["path"][-2] if len(pe["path"]) >= 2 else ""
        label  = f"PoKorb_{parent_app}" if parent_app else "PoKorb"
        parent = f"Pe_{parent_app}"     if parent_app else "Pe"

        values = [
            label,
            f"Postkorb {parent_app}" if parent_app else "Postkorb",
            parent,
            "Posteingang",
            "",
            "",
        ]
        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = LEFT; c.border = BOX2
        ws.cell(row=r, column=1).fill = YELLOW
        for j in range(7, 12):
            c = ws.cell(row=r, column=j, value="")
            c.alignment = LEFT; c.border = BOX2
        ws.cell(row=r, column=1).font = BLACK
        for j in range(2, 12):
            ws.cell(row=r, column=j).font = BLACK

        if pe.get("disabled"):
            gray_out_row(ws, r, 1, 11)

        r += 1

    ws.freeze_panes = "A2"
    autosize_columns(ws, min_w=10, max_w=120)
    return ws

# ---------------------------------------------------------------------------
# Sheet 3: Geschäftsrollen (with spacer column & borders)
# ---------------------------------------------------------------------------

def get_label(node):
    return (node.get("appName") or node.get("name") or "").strip()

def get_desc(node):
    return (node.get("description") or "").strip()

def is_ablagen_label(lbl: str) -> bool:
    l = (lbl or "").lower()
    return l.startswith("ab_") or l.startswith("pe_") or l.startswith("sb_") or l.startswith("sb-") or l.startswith("sb-thema")

def is_ltg_or_allg_for(lbl: str, base: str) -> bool:
    if "_" not in lbl:
        return False
    prefix, suffix = lbl.split("_", 1)
    return (prefix in {"Ltg", "Allg"}) and (suffix == base)

def preorder_nodes(node):
    yield node
    for ch in node.get("children") or []:
        yield from preorder_nodes(ch)

def collect_ablagen_nodes_excluding_ab_org(children, org_label: str):
    """
    If there is an 'Ab_{org}' root, skip that node itself and include its descendants;
    otherwise include any ablagen-like subtree roots.
    """
    out = []
    skip_label = f"Ab_{org_label}"
    for ch in children or []:
        lbl = get_label(ch)
        if lbl.lower().startswith("ab_") and lbl == skip_label:
            for i, n in enumerate(preorder_nodes(ch)):
                if i == 0:
                    continue
                out.append(n)
        elif is_ablagen_label(lbl):
            out.extend(list(preorder_nodes(ch)))
    return out

def fill_row(ws, row, start_col, end_col, fill):
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = fill

def roles_sheet_header(ws, roles_count: int):
    """
    Header with spacer column:
      col1: Ablagen/Hierarchieelemente
      col2: spacer
      col3: Beschreibung
      col4..: Role blocks (Lesen | Schreiben | LA) × roles_count
    """
    h1 = ws.cell(row=1, column=1, value="Ablagen / Hierarchieelemente"); h1.font = BLACKB; h1.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)

    # spacer col 2 (empty on purpose)

    h2 = ws.cell(row=1, column=3, value="Beschreibung"); h2.font = BLACKB; h2.alignment = LEFT
    ws.merge_cells(start_row=1, start_column=3, end_row=2, end_column=3)

    col = 4
    for k in range(roles_count):
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
        head = ws.cell(row=1, column=col, value=f"{ROLE_HEADER_TEXT} {k+1}")
        head.font = BLACKB; head.alignment = CENTER
        ws.cell(row=2, column=col+0, value="Lesen").font = BLACKB; ws.cell(row=2, column=col+0).alignment = CENTER
        ws.cell(row=2, column=col+1, value="Schreiben").font = BLACKB; ws.cell(row=2, column=col+1).alignment = CENTER
        ws.cell(row=2, column=col+2, value="LA").font = BLACKB; ws.cell(row=2, column=col+2).alignment = CENTER
        col += 3

    ws.column_dimensions[get_column_letter(1)].width = 38
    ws.column_dimensions[get_column_letter(2)].width = 2   # spacer
    ws.column_dimensions[get_column_letter(3)].width = 56
    for c in range(4, 4 + roles_count * 3):
        ws.column_dimensions[get_column_letter(c)].width = 10

def apply_role_vertical_borders(ws, roles_count: int, end_row: int):
    """
    Vertical separators:
      - thin gray between Lesen | Schreiben | LA
      - thick after each role group
      - thin gray at the left of Beschreibung (col 3)
      - thick at far right
    """
    total_cols = 3 + roles_count * 3

    for k in range(roles_count):
        base = 4 + k * 3
        mid  = base + 1
        last = base + 2

        for r in range(1, end_row + 1):
            c1 = ws.cell(row=r, column=base)
            c1.border = Border(left=c1.border.left, right=GRID_GRAY, top=c1.border.top, bottom=c1.border.bottom)

            c2 = ws.cell(row=r, column=mid)
            c2.border = Border(left=c2.border.left, right=GRID_GRAY, top=c2.border.top, bottom=c2.border.bottom)

            c3 = ws.cell(row=r, column=last)
            c3.border = Border(left=c3.border.left, right=THICK, top=c3.border.top, bottom=c3.border.bottom)

    for r in range(1, end_row + 1):
        desc_cell = ws.cell(row=r, column=3)
        b = desc_cell.border
        desc_cell.border = Border(left=GRID_GRAY, right=b.right, top=b.top, bottom=b.bottom)

    for r in range(1, end_row + 1):
        cell = ws.cell(row=r, column=total_cols)
        cell.border = Border(left=cell.border.left, right=THICK, top=cell.border.top, bottom=cell.border.bottom)

def apply_gray_horizontal_grid(ws, start_row: int, end_row: int, total_cols: int):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=GRID_GRAY)
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(left=b.left, right=b.right, top=GRID_GRAY, bottom=b.bottom)

def apply_thick_bottom(ws, total_cols: int, end_row: int):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=end_row, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK_BOTTOM)

def gray_out_row_full(ws, row, total_cols):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=row, column=c)
        bold = bool(cell.font and cell.font.bold)
        cell.fill = DISABLED_FILL
        cell.font = GRAYB if bold else GRAY

def write_role_row(ws, row_idx, label, desc, total_cols, fill_full_row=None):
    ws.cell(row=row_idx, column=1, value=label).alignment = LEFT
    ws.cell(row=row_idx, column=3, value=desc).alignment = LEFT
    if fill_full_row:
        fill_row(ws, row_idx, 1, total_cols, fill_full_row)

def get_label(node):
    return (node.get("appName") or node.get("name") or "").strip()

def get_desc(node):
    return (node.get("description") or "").strip()

def is_ablagen_label(lbl: str) -> bool:
    l = (lbl or "").lower()
    return l.startswith("ab_") or l.startswith("pe_") or l.startswith("sb_") or l.startswith("sb-") or l.startswith("sb-thema")

def is_ltg_or_allg_for(lbl: str, base: str) -> bool:
    if "_" not in lbl:
        return False
    prefix, suffix = lbl.split("_", 1)
    return (prefix in {"Ltg", "Allg"}) and (suffix == base)

def preorder_nodes(node):
    yield node
    for ch in node.get("children") or []:
        yield from preorder_nodes(ch)

def collect_ablagen_nodes_excluding_ab_org(children, org_label: str):
    out = []
    skip_label = f"Ab_{org_label}"
    for ch in children or []:
        lbl = get_label(ch)
        if lbl.lower().startswith("ab_") and lbl == skip_label:
            for i, n in enumerate(preorder_nodes(ch)):
                if i == 0:
                    continue
                out.append(n)
        elif is_ablagen_label(lbl):
            out.extend(list(preorder_nodes(ch)))
    return out

def write_roles_for_node(ws, node, row_idx, roles_count):
    total_cols = 3 + roles_count * 3
    org_label = get_label(node)
    if not org_label:
        return row_idx

    children = node.get("children") or []

    # Org section header
    write_role_row(ws, row_idx, org_label, get_desc(node), total_cols, fill_full_row=TITLE_GRAY)
    if node.get("enabled") is False:
        gray_out_row_full(ws, row_idx, total_cols)
    row_idx += 1

    # Ltg / Allg under the org
    ltg_allg_nodes = [ch for ch in children if is_ltg_or_allg_for(get_label(ch), org_label)]
    for ch in ltg_allg_nodes:
        write_role_row(ws, row_idx, get_label(ch), get_desc(ch), total_cols, fill_full_row=PALEGR)
        if ch.get("enabled") is False:
            gray_out_row_full(ws, row_idx, total_cols)
        row_idx += 1

    # Spacer
    row_idx += 1

    # Section header for Ab_{Org}
    write_role_row(ws, row_idx, f"Ab_{org_label}", "", total_cols, fill_full_row=TITLE_GRAY)
    row_idx += 1

    # Ablagen subtree rows
    ablagen_nodes = collect_ablagen_nodes_excluding_ab_org(children, org_label)
    for n in ablagen_nodes:
        write_role_row(ws, row_idx, get_label(n), get_desc(n), total_cols, fill_full_row=PALEGR)
        if n.get("enabled") is False:
            gray_out_row_full(ws, row_idx, total_cols)
        row_idx += 1

    # Spacer after the ablagen block
    row_idx += 1

    # Recurse into remaining children
    ablagen_roots = {get_label(ch) for ch in children if is_ablagen_label(get_label(ch))}
    ltg_allg_set  = {get_label(ch) for ch in ltg_allg_nodes}
    for ch in children:
        lbl = get_label(ch)
        if lbl in ltg_allg_set or lbl in ablagen_roots:
            continue
        row_idx = write_roles_for_node(ws, ch, row_idx, roles_count)

    return row_idx

def find_last_populated_row(ws, total_cols: int) -> int:
    """Find last data row by checking cols 1 or 3 (name/description)."""
    max_row = ws.max_row
    for r in range(max_row, 2, -1):
        if (ws.cell(row=r, column=1).value not in (None, "")
            or ws.cell(row=r, column=3).value not in (None, "")):
            return r
    return 2

def add_third_sheet(wb: Workbook, tree, roles_count:int):
    ws = wb.create_sheet(title=SHEET3_NAME)
    roles_sheet_header(ws, roles_count)

    working_nodes = strip_prefix_levels(tree, SKIP_PARENTS)

    r = 3
    for idx, top in enumerate(working_nodes):
        r = write_roles_for_node(ws, top, r, roles_count)
        if idx < len(working_nodes) - 1:
            r += 1  # blank line BETWEEN top-level groups only

    total_cols = 3 + roles_count * 3
    end_row = find_last_populated_row(ws, total_cols)

    apply_role_vertical_borders(ws, roles_count, end_row)
    apply_gray_horizontal_grid(ws, start_row=3, end_row=end_row, total_cols=total_cols)
    apply_thick_bottom(ws, total_cols, end_row)

    ws.freeze_panes = "D3"
    autosize_columns(ws, min_w=8, max_w=60)
    return ws

# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

@app.post("/generate-excel")
async def generate_excel(request: Request):
    """
    JSON body:
      {
        "tree": [...],                 # REQUIRED (wrapped-or-unwrapped is fine)
        "sheets": ["GE","Ablage","Roles"],   # OPTIONAL
        "rolesCount": 10              # OPTIONAL (Roles sheet placeholder count)
      }
    """
    data = await request.json()
    tree = data.get("tree")
    if not tree:
        return {"error": "Missing 'tree' in JSON body"}

    sheets = data.get("sheets") or ["GE", "Ablage", "Roles"]
    roles_count = int(data.get("rolesCount") or ROLES_COUNT)

    max_depth = max(0, tree_max_depth(tree))
    last_name_col = max_depth + 1

    if "GE" in sheets:
        wb, ws, cols = create_sheet(
            last_name_col,
            perm1_cols=len(LEFT_HEADERS),
            perm2_cols=len(PERM_HEADERS),
            max_depth=max_depth
        )
        write_rows(
            ws, tree, DATA_START_ROW, level=0, last_stack=[],
            cols=cols, lineage_names=[], lineage_apps=[]
        )
        autosize_columns(ws, min_w=8, max_w=60)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Ablage" in sheets:
        add_second_sheet(wb, tree)

    if "Roles" in sheets:
        add_third_sheet(wb, tree, roles_count)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tree.xlsx"}
    )
