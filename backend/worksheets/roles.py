# sheet_roles.py

from openpyxl import Workbook
from openpyxl.styles import Border

from config import (
    SHEET3_NAME,
    SKIP_PARENTS,
    ROLE_HEADER_TEXT,
)
from styles import (
    TITLE_GRAY, PALEGR, DISABLED_FILL,
    GRID_GRAY, THICK, THICK_BOTTOM,
    BOX, LEFT, BLACKB, BLACK, GRAY, GRAYB,
)
from utils import strip_prefix_levels, autosize_columns


def fill_row(ws, row, start_col, end_col, fill):
    for c in range(start_col, end_col + 1):
        ws.cell(row=row, column=c).fill = fill


def roles_sheet_header(ws, roles_count: int):
    ws.cell(row=2, column=1, value="Ablagen / Hierarchieelemente").font = BLACKB
    ws.cell(row=2, column=1).alignment = LEFT

    ws.cell(row=2, column=3, value="Beschreibung").font = BLACKB
    ws.cell(row=2, column=3).alignment = LEFT

    col = 4
    for k in range(roles_count):
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
        head = ws.cell(row=1, column=col, value=f"{ROLE_HEADER_TEXT} {k+1}")
        head.font = BLACKB
        head.alignment = LEFT

        ws.cell(row=2, column=col+0, value="Lesen").font = BLACKB
        ws.cell(row=2, column=col+0).alignment = LEFT
        ws.cell(row=2, column=col+1, value="Schreiben").font = BLACKB
        ws.cell(row=2, column=col+1).alignment = LEFT
        ws.cell(row=2, column=col+2, value="LA").font = BLACKB
        ws.cell(row=2, column=col+2).alignment = LEFT
        col += 3

    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(1)].width = 38
    ws.column_dimensions[get_column_letter(2)].width = 2   # spacer
    ws.column_dimensions[get_column_letter(3)].width = 56
    for c in range(4, 4 + roles_count * 3):
        ws.column_dimensions[get_column_letter(c)].width = 10


def apply_role_vertical_borders(ws, roles_count: int, end_row: int):
    total_cols = 3 + roles_count * 3

    for k in range(roles_count):
        base = 4 + k * 3
        mid  = base + 1
        last = base + 2

        for r in range(1, end_row + 1):
            c1 = ws.cell(row=r, column=base)
            c1.border = Border(
                left=c1.border.left,
                right=GRID_GRAY,
                top=c1.border.top,
                bottom=c1.border.bottom,
            )

            c2 = ws.cell(row=r, column=mid)
            c2.border = Border(
                left=c2.border.left,
                right=GRID_GRAY,
                top=c2.border.top,
                bottom=c2.border.bottom,
            )

            c3 = ws.cell(row=r, column=last)
            c3.border = Border(
                left=c3.border.left,
                right=THICK,
                top=c3.border.top,
                bottom=c3.border.bottom,
            )

    for r in range(1, end_row + 1):
        c = ws.cell(row=r, column=4)
        b = c.border
        c.border = Border(left=THICK, right=b.right, top=b.top, bottom=b.bottom)

    for r in range(1, end_row + 1):
        sp = ws.cell(row=r, column=2)
        b = sp.border
        sp.border = Border(left=GRID_GRAY, right=b.right, top=b.top, bottom=b.bottom)

    for r in range(1, end_row + 1):
        desc_cell = ws.cell(row=r, column=3)
        b = desc_cell.border
        desc_cell.border = Border(left=GRID_GRAY, right=b.right, top=b.top, bottom=b.bottom)

    for r in range(1, end_row + 1):
        cell = ws.cell(row=r, column=total_cols)
        cell.border = Border(
            left=cell.border.left,
            right=THICK,
            top=cell.border.top,
            bottom=cell.border.bottom,
        )


def apply_gray_horizontal_grid(ws, start_row: int, end_row: int, total_cols: int):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=2, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=GRID_GRAY)
    for r in range(start_row, end_row + 1):
        for c in range(1, total_cols + 1):
            cell = ws.cell(row=r, column=c)
            b = cell.border
            cell.border = Border(left=b.left, right=b.right, top=GRID_GRAY, bottom=b.bottom)


def apply_thick_bottom(ws, total_cols: int, end_row: int):
    for c in range(1, total_cols + 1):
        cell = ws.cell(row=end_row, column=c)
        b = cell.border
        cell.border = Border(left=b.left, right=b.right, top=b.top, bottom=THICK_BOTTOM)


def write_role_row(ws, row_idx, label, desc, total_cols, fill_full_row=None):
    ws.cell(row=row_idx, column=1, value=label).alignment = LEFT
    ws.cell(row=row_idx, column=3, value=desc).alignment = LEFT
    if fill_full_row:
        fill_row(ws, row_idx, 1, total_cols, fill_full_row)


def get_label(node):
    return (node.get("appName") or node.get("name") or "").strip()


def get_desc(node):
    return (node.get("description") or "").strip()


def is_ablagen_label(lbl: str) -> bool:
    l = (lbl or "").lower()
    return (
        l.startswith("ab_") or
        l.startswith("pe_") or
        l.startswith("sb_") or
        l.startswith("sb-") or
        l.startswith("sb-thema")
    )


def is_ltg_or_allg_for(lbl: str, base: str) -> bool:
    if "_" not in lbl:
        return False
    prefix, suffix = lbl.split("_", 1)
    return (prefix in {"Ltg", "Allg"}) and (suffix == base)


def preorder_nodes(node):
    yield node
    for ch in node.get("children") or []:
        yield from preorder_nodes(ch)


def collect_ablagen_nodes_excluding_ab_org(children, org_label: str):
    out = []
    skip_label = f"Ab_{org_label}"
    for ch in children or []:
        lbl = get_label(ch)
        if lbl.lower().startswith("ab_") and lbl == skip_label:
            for i, n in enumerate(preorder_nodes(ch)):
                if i == 0:
                    continue
                out.append(n)
        elif is_ablagen_label(lbl):
            out.extend(list(preorder_nodes(ch)))
    return out


def write_roles_for_node(ws, node, row_idx, roles_count):
    total_cols = 3 + roles_count * 3
    org_label = get_label(node)
    if not org_label:
        return row_idx

    children = node.get("children") or []

    write_role_row(ws, row_idx, org_label, get_desc(node), total_cols, fill_full_row=TITLE_GRAY)
    if node.get("enabled") is False:
        from utils import gray_out_row_full
        gray_out_row_full(ws, row_idx, total_cols)
    row_idx += 1

    ltg_allg_nodes = [ch for ch in children if is_ltg_or_allg_for(get_label(ch), org_label)]
    for ch in ltg_allg_nodes:
        write_role_row(ws, row_idx, get_label(ch), get_desc(ch), total_cols, fill_full_row=PALEGR)
        if ch.get("enabled") is False:
            from utils import gray_out_row_full
            gray_out_row_full(ws, row_idx, total_cols)
        row_idx += 1

    row_idx += 1

    write_role_row(ws, row_idx, f"Ab_{org_label}", "", total_cols, fill_full_row=TITLE_GRAY)
    row_idx += 1

    ablagen_nodes = collect_ablagen_nodes_excluding_ab_org(children, org_label)
    for n in ablagen_nodes:
        write_role_row(ws, row_idx, get_label(n), get_desc(n), total_cols, fill_full_row=PALEGR)
        if n.get("enabled") is False:
            from utils import gray_out_row_full
            gray_out_row_full(ws, row_idx, total_cols)
        row_idx += 1

    row_idx += 1

    ablagen_roots = {get_label(ch) for ch in children if is_ablagen_label(get_label(ch))}
    ltg_allg_set  = {get_label(ch) for ch in ltg_allg_nodes}
    for ch in children:
        lbl = get_label(ch)
        if lbl in ltg_allg_set or lbl in ablagen_roots:
            continue
        row_idx = write_roles_for_node(ws, ch, row_idx, roles_count)

    return row_idx


def find_last_populated_row(ws, total_cols: int) -> int:
    max_row = ws.max_row
    for r in range(max_row, 2, -1):
        if (ws.cell(row=r, column=1).value not in (None, "")
            or ws.cell(row=r, column=3).value not in (None, "")):
            return r
    return 2


def add_third_sheet(wb: Workbook, tree, roles_count:int):
    ws = wb.create_sheet(title=SHEET3_NAME)
    roles_sheet_header(ws, roles_count)

    working_nodes = strip_prefix_levels(tree, SKIP_PARENTS)

    r = 3
    for idx, top in enumerate(working_nodes):
        r = write_roles_for_node(ws, top, r, roles_count)
        if idx < len(working_nodes) - 1:
            r += 1

    total_cols = 3 + roles_count * 3
    end_row = find_last_populated_row(ws, total_cols)

    apply_role_vertical_borders(ws, roles_count, end_row)
    apply_gray_horizontal_grid(ws, start_row=3, end_row=end_row, total_cols=total_cols)
    apply_thick_bottom(ws, total_cols, end_row)

    USER_ROWS = 15
    bottom_start = end_row + 1
    bottom_end = end_row + USER_ROWS

    ws.merge_cells(start_row=bottom_start, start_column=1, end_row=bottom_end, end_column=3)
    label_text = (
        "Benutzer\n\n"
        "Pro Zelle genau EIN Benutzer.\n"
        "Benutzer nur in diesem Bereich eintragen – je Rolle in den zugehörigen Feldern.\n"
        "Eintragung spaltenweise oder zeilenweise möglich."
    )
    label_cell = ws.cell(row=bottom_start, column=1, value=label_text)
    label_cell.font = BLACKB
    label_cell.alignment = LEFT
    label_cell.border = BOX

    for r_idx in range(bottom_start, bottom_end + 1):
        for c_idx in range(4, total_cols + 1):
            c = ws.cell(row=r_idx, column=c_idx)
            if c.value is None:
                c.value = ""

    apply_role_vertical_borders(ws, roles_count, bottom_end)
    apply_gray_horizontal_grid(ws, start_row=bottom_start, end_row=bottom_end, total_cols=total_cols)
    apply_thick_bottom(ws, total_cols, bottom_end)

    ws.freeze_panes = "D3"
    autosize_columns(ws, min_w=8, max_w=60)
    return ws
