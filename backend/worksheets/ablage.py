from openpyxl import Workbook
from openpyxl.styles import PatternFill

from config import (
    SHEET2_NAME,
    SKIP_PARENTS,
    ORG_NAME,
    ADMIN_ACCOUNT,
    PREFIX,
)
from styles import (
    PALEGR,
    BLACKB,
    BLACK,
    LEFT,
    BOX2,
    BLUE,
    WHITEB,
    GRAYB,
    THICK_TOP,
    THICK_BOTTOM,
)

from utils import strip_prefix_levels, gray_out_row, autosize_columns, norm_token


USE_THICK_TOP = False
USE_THICK_BOTTOM = False
USE_GROUP_BOTTOM = False
USE_BLANK_BEFORE_PARENT = True

DISABLED_BLUE = PatternFill("solid", fgColor="9FB7D9")
POKORB_FILL = PatternFill("solid", fgColor="FFF2CC")


def allowed_types_for(label, typ):
    lab = (label or "").lower()
    if typ == "Posteingang" or lab.startswith("pe_") or "poeing" in lab:
        return "GOV_WORKING_FOLDER_INBOX"
    return "GOV_FILE" if typ == "Aktenablage" else ""


def break_inheritance_from_node(node, desc_text):
    v = node.get("unterbrechen")
    if isinstance(v, str) and v.strip().lower() in {"wahr", "true", "ja"}:
        return "WAHR"
    if v is True:
        return "WAHR"
    if "ohne leitungszugriff" in (desc_text or "").lower():
        return "WAHR"
    return ""


def build_group_key_from_names(path_names):
    tokens = [norm_token(x) for x in path_names if (x or "").strip()]
    return PREFIX + "_".join(tokens) if tokens else ""


def set_row_top_thick(ws, row, col_start=1, col_end=11):
    if not USE_THICK_TOP:
        return
    from openpyxl.styles import Border, Side
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(
            left=b.left or Side(style="thin", color="000000"),
            right=b.right or Side(style="thin", color="000000"),
            top=THICK_TOP,
            bottom=b.bottom or Side(style="thin", color="000000"),
        )


def set_row_bottom_thick(ws, row, col_start=1, col_end=11):
    if not USE_THICK_BOTTOM:
        return
    from openpyxl.styles import Border, Side
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        b = cell.border
        cell.border = Border(
            left=b.left or Side(style="thin", color="000000"),
            right=b.right or Side(style="thin", color="000000"),
            top=b.top or Side(style="thin", color="000000"),
            bottom=THICK_BOTTOM,
        )


def set_group_bottom_thick(ws, row_last, col_start=1, col_end=11):
    if USE_GROUP_BOTTOM:
        set_row_bottom_thick(ws, row_last, col_start, col_end)


def make_addr(group_key: str, suffix: str = "") -> str:
    key = f"{group_key}-{suffix}" if suffix else group_key
    return f"{key}@{ORG_NAME}"


def is_poe_label(label: str) -> bool:
    low = (label or "").lower()
    return low.startswith("pe_") or "poeing" in low


def write_group_recursive(ws, node, r, path_names, path_apps, poeings):
    name = (node.get("name") or "").strip()
    appn = (node.get("appName") or name).strip()
    kids = node.get("children") or []
    desc = (node.get("description") or "").strip()
    label = appn if appn else (name if name else "(unnamed)")
    is_parent = bool(kids)
    disabled = node.get("enabled") is False
    parent_raw = path_apps[-1] if path_apps else ""

    if (
        USE_BLANK_BEFORE_PARENT
        and is_parent
        and (path_names or path_apps)
    ):
        r += 1

    display_a = label
    parent_display = parent_raw
    typ = "Hierarchieelement" if is_parent else "Aktenablage"

    full_names_path = path_names + [name]
    perm_key = build_group_key_from_names(full_names_path)
    ancestor_keys = [
        build_group_key_from_names(full_names_path[:i])
        for i in range(1, len(full_names_path) + 1)
    ]

    values = [
        display_a,
        desc,
        parent_display,
        typ,
        break_inheritance_from_node(node, desc),
        allowed_types_for(label, "Posteingang" if is_poe_label(label) else typ),
    ]

    for j, v in enumerate(values, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.alignment = LEFT
        c.border = BOX2

    g_list = [make_addr(perm_key, "RO")]
    h_list = [make_addr(perm_key, "")]
    i_list = [make_addr(perm_key, "FA")]
    j_list_vals = [make_addr(k, "LA") for k in ancestor_keys]
    # CHANGED: Ablageadmin now also uses all ancestor keys (like Löschadmin)
    k_list = [make_addr(k, "AA") for k in ancestor_keys]

    for offset, items in enumerate([g_list, h_list, i_list, j_list_vals, k_list], start=7):
        cell_val = ";".join(items) + f";{ADMIN_ACCOUNT}"
        c = ws.cell(row=r, column=offset, value=cell_val)
        c.alignment = LEFT
        c.border = BOX2

    if is_parent:
        fill = DISABLED_BLUE if disabled else BLUE
        font = GRAYB if disabled else WHITEB
    else:
        fill = PALEGR
        font = GRAYB if disabled else BLACK

    for col in range(1, 12):
        ws.cell(row=r, column=col).fill = fill
        ws.cell(row=r, column=col).font = font

    if not is_parent and disabled:
        gray_out_row(ws, r, 1, 11)

    if is_parent:
        set_row_top_thick(ws, r, 1, 11)

    cur = r

    if is_poe_label(label):
        poeings.append({"path": list(path_apps), "disabled": disabled})

    for child in kids:
        cur += 1
        cur, _ = write_group_recursive(
            ws,
            child,
            cur,
            path_names=full_names_path,
            path_apps=path_apps + [label],
            poeings=poeings,
        )

    if is_parent:
        set_group_bottom_thick(ws, cur, 1, 11)

    return cur, poeings


def add_second_sheet(wb: Workbook, tree):
    ws = wb.create_sheet(title=SHEET2_NAME)

    headers = [
        "Bezeichnung strukturierte Ablage",
        "Beschreibung",
        "Eltern strukturierte Ablage",
        "Art",
        "Vererbung aus übergeordnetem Element unterbrechen",
        "Erlaubte Aktentypen",
        "Lesen",
        "Schreiben",
        "Administrieren",
        "Löschadministration",
        "Ablageadministration",
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = BLACKB
        c.alignment = LEFT
        c.border = BOX2

    for col_idx in range(7, 12):
        ws.cell(row=1, column=col_idx).fill = PALEGR

    working_nodes = strip_prefix_levels(tree, SKIP_PARENTS)

    r = 2
    all_poeings = []

    for top in working_nodes:
        r, poe = write_group_recursive(ws, top, r, [], [], [])
        all_poeings.extend(poe)
        r += 1

    first_pokorb_row = None

    for pe in all_poeings:
        parent_app = pe["path"][-2] if len(pe["path"]) >= 2 else ""
        label = f"PoKorb_Pe_{parent_app}" if parent_app else "PoKorb"
        parent = f"Pe_{parent_app}" if parent_app else "Pe"

        values = [
            label,
            f"Postkorb {parent_app}" if parent_app else "Postkorb",
            parent,
            "Posteingang",
            "",
            "",
        ]

        for j, v in enumerate(values, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.alignment = LEFT
            c.border = BOX2

        disabled = pe.get("disabled")
        fill = DISABLED_BLUE if disabled else POKORB_FILL
        font = GRAYB if disabled else BLACK

        for col in range(1, 12):
            cell = ws.cell(row=r, column=col)
            cell.fill = fill
            cell.font = font
            cell.border = BOX2

        if first_pokorb_row is None:
            first_pokorb_row = r
            from openpyxl.styles import Border, Side
            thin_top = Side(style="medium", color="000000")
            for col in range(1, 12):
                cell = ws.cell(row=first_pokorb_row, column=col)
                b = cell.border
                cell.border = Border(
                    left=b.left,
                    right=b.right,
                    top=thin_top,
                    bottom=b.bottom,
                )

        if disabled:
            gray_out_row(ws, r, 1, 11)

        r += 1

    ws.freeze_panes = "A2"
    autosize_columns(ws, min_w=10, max_w=120)
    return ws
