from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import (
    SHEET_NAME,
    GROUPS_FROM_LEVEL,
    PREFIX,
    PERM_HEADERS,
    PERM_SUFFIX,
    LEFT_HEADERS,
    LEFT_SUFFIX,
    FORCE_BLUE,
    ROW_BAND,
    ROW_TITLE,
    ROW_CAPTION,
    ROW_HEADERS,
)
from styles import (
    ORANGE, CYAN, PALEGR, BLUE, LIME, TITLE_GRAY,
    WHITEB, BLACKB, GRAY, LEFT, CENTER, BOX,
)
from utils import norm_token, gray_out_row


def create_sheet(last_name_col: int, perm1_cols: int, perm2_cols: int, max_depth: int):
    spacer1 = last_name_col + 1
    perm1_s = spacer1 + 1
    perm1_e = perm1_s + perm1_cols - 1
    spacer2 = perm1_e + 1
    perm2_s = spacer2 + 1
    perm2_e = perm2_s + perm2_cols - 1
    spacer3 = perm2_e + 1
    flat_col = spacer3 + 1
    spacer4 = flat_col + 1
    tree_base = spacer4 + 1
    tree_end  = tree_base + max_depth

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for c in range(1, last_name_col):
        ws.column_dimensions[get_column_letter(c)].width = 4
    ws.column_dimensions[get_column_letter(last_name_col)].width = 26
    for c in [spacer1, spacer2, spacer3, spacer4]:
        ws.column_dimensions[get_column_letter(c)].width = 2
    for c in list(range(perm1_s, perm1_e+1)) + list(range(perm2_s, perm2_e+1)):
        ws.column_dimensions[get_column_letter(c)].width = 22
    ws.column_dimensions[get_column_letter(flat_col)].width = 26
    for c in range(tree_base, tree_end+1):
        ws.column_dimensions[get_column_letter(c)].width = 6

    for r,h in [(ROW_BAND,24),(ROW_TITLE,20),(ROW_HEADERS-1,18)]:
        ws.row_dimensions[r].height = h

    ws.merge_cells(start_row=ROW_BAND, start_column=perm1_s, end_row=ROW_BAND, end_column=perm1_e)
    c1 = ws.cell(row=ROW_BAND, column=perm1_s, value="eDir")
    c1.fill = ORANGE; c1.font = WHITEB; c1.alignment = CENTER; c1.border = BOX

    chip = ws.cell(row=ROW_BAND, column=perm2_s, value="eDir")
    chip.fill = ORANGE; chip.font = WHITEB; chip.alignment = CENTER; chip.border = BOX

    ws.merge_cells(start_row=ROW_BAND, start_column=perm2_s+1, end_row=ROW_BAND, end_column=perm2_e)
    c2 = ws.cell(row=ROW_BAND, column=perm2_s+1, value="AD(DFSW)")
    c2.fill = CYAN; c2.font = BLACKB; c2.alignment = CENTER; c2.border = BOX

    ws.cell(row=ROW_TITLE, column=1, value="Struktur Gruppen mit Schreibzugriff").font = BLACKB
    ws.merge_cells(start_row=ROW_CAPTION, start_column=perm1_s, end_row=ROW_CAPTION, end_column=perm1_e)
    ws.cell(row=ROW_CAPTION, column=perm1_s, value="auf Knoten zusätzlich anzulegende Gruppen").font = BLACKB

    for j,hdr in enumerate(LEFT_HEADERS, start=perm1_s):
        hc = ws.cell(row=ROW_HEADERS, column=j, value=hdr)
        hc.font = BLACKB; hc.alignment = CENTER; hc.border = BOX
    for j,hdr in enumerate(PERM_HEADERS, start=perm2_s):
        hc = ws.cell(row=ROW_HEADERS, column=j, value=hdr)
        hc.font = BLACKB; hc.alignment = CENTER; hc.border = BOX

    ws.merge_cells(start_row=ROW_BAND, start_column=flat_col, end_row=ROW_BAND, end_column=tree_end)
    ab = ws.cell(row=ROW_BAND, column=flat_col, value="nscale strukturierte Ablage")
    ab.fill = PALEGR; ab.font = BLACKB; ab.alignment = LEFT; ab.border = BOX

    ws.cell(row=ROW_HEADERS-1, column=flat_col, value="Liste").font = BLACKB
    ws.merge_cells(start_row=ROW_HEADERS-1, start_column=tree_base, end_row=ROW_HEADERS-1, end_column=tree_end)
    bh = ws.cell(row=ROW_HEADERS-1, column=tree_base, value="Baum")
    bh.font = BLACKB; bh.alignment = LEFT

    cols = {
        "spacer1": spacer1, "perm1_s": perm1_s, "perm1_e": perm1_e,
        "spacer2": spacer2, "perm2_s": perm2_s, "perm2_e": perm2_e,
        "spacer3": spacer3, "flat_col": flat_col, "spacer4": spacer4,
        "tree_base": tree_base, "tree_end": tree_end, "max_depth": max_depth
    }
    return wb, ws, cols


# ASCII-tree helpers

def draw_verticals_before_elbow(ws, row: int, level: int, last_stack, base_col: int = 1):
    if level <= 1:
        return
    for d in range(0, level - 1):
        idx = d + 1
        if idx < len(last_stack) and not last_stack[idx]:
            c = ws.cell(row=row, column=base_col + d, value="│")
            c.alignment = CENTER
            c.font = GRAY


def draw_verticals_through_subtree(ws, row: int, level: int, last_stack, base_col: int = 1):
    if level <= 0:
        return
    for d in range(0, level):
        idx = d + 1
        if idx < len(last_stack) and not last_stack[idx]:
            c = ws.cell(row=row, column=base_col + d, value="│")
            c.alignment = CENTER
            c.font = GRAY


def draw_connectors(ws, row: int, level: int, last_stack, base_col: int = 1):
    if level <= 0:
        return
    draw_verticals_before_elbow(ws, row, level, last_stack, base_col=base_col)
    elbow_col = base_col + (level - 1)
    elbow = "└" if last_stack[-1] else "├"
    c = ws.cell(row=row, column=elbow_col, value=elbow)
    c.alignment = CENTER
    c.font = GRAY


def extend_connectors(ws, start_row: int, end_row: int, level: int, last_stack, base_col: int = 1):
    if end_row < start_row or level < 0:
        return
    for r in range(start_row, end_row + 1):
        draw_verticals_through_subtree(ws, r, level, last_stack, base_col=base_col)


def style_cell_like_node(cell, label: str, is_container: bool):
    cell.alignment = LEFT
    cell.border = BOX
    if is_container or label in FORCE_BLUE:
        cell.fill = BLUE
        cell.font = WHITEB
    else:
        cell.fill = LIME
        cell.font = BLACKB


def style_name_cell(cell, node_name: str, has_children: bool):
    cell.alignment = LEFT
    cell.border = BOX
    if has_children or node_name in FORCE_BLUE:
        cell.fill = BLUE
        cell.font = WHITEB
    else:
        cell.fill = LIME
        cell.font = BLACKB


def is_ablg(s: str) -> bool:
    return (s or "").strip().lower() == "ablgoe"


def is_poeing(s: str) -> bool:
    return (s or "").strip().lower() == "poeing"


def write_rows(ws, nodes, row, level, last_stack, cols, lineage_names, lineage_apps):
    name_col = level + 1
    for i, node in enumerate(nodes):
        name = (node.get("name") or "").strip()
        appn = (node.get("appName") or name).strip()
        children = node.get("children") or []
        if not name and not children:
            continue

        is_last = (i == len(nodes) - 1)
        disabled = (node.get("enabled") is False)

        draw_connectors(ws, row, level, last_stack + [is_last], base_col=1)

        nc = ws.cell(row=row, column=name_col, value=name if name else "(unnamed)")
        style_name_cell(nc, name, bool(children))

        for dc in range(name_col + 1, cols["spacer1"]):
            d = ws.cell(row=row, column=dc, value="-")
            d.alignment = CENTER
            d.border = BOX

        for sc in [cols["spacer1"], cols["spacer2"], cols["spacer3"], cols["spacer4"]]:
            ws.cell(row=row, column=sc, value="")

        if level >= GROUPS_FROM_LEVEL and name:
            for j, suf in enumerate(LEFT_SUFFIX, start=cols["perm1_s"]):
                val = f"{name}-{suf}" if suf else name
                g = ws.cell(row=row, column=j, value=val)
                g.alignment = LEFT
                g.border = BOX

        if level >= GROUPS_FROM_LEVEL and name:
            tokens = lineage_names + [norm_token(name)]
            start = min(GROUPS_FROM_LEVEL, len(tokens) - 1)
            key = PREFIX + "_".join(tokens[start:])
            for j, suf in enumerate(PERM_SUFFIX, start=cols["perm2_s"]):
                val = f"{key}-{suf}" if suf else key
                g = ws.cell(row=row, column=j, value=val)
                g.alignment = LEFT
                g.border = BOX

        flat_label = appn if appn else (name if name else "(unnamed)")
        fc = ws.cell(row=row, column=cols["flat_col"], value=flat_label)
        style_cell_like_node(fc, flat_label, bool(children))

        r_name_col = cols["tree_base"] + level
        tv_label = appn if appn else (name if name else "(unnamed)")
        tv = ws.cell(row=row, column=r_name_col, value=tv_label)
        style_cell_like_node(tv, tv_label, bool(children))

        if disabled:
            gray_out_row(ws, row, 1, cols["tree_end"])

        subtree_start_row = row
        row += 1

        if children:
            row = write_rows(
                ws, children, row, level + 1, last_stack + [is_last], cols,
                lineage_names + [norm_token(name)],
                lineage_apps + [appn]
            )

        subtree_end_row = row - 1
        extend_connectors(ws, subtree_start_row + 1, subtree_end_row, level, last_stack + [is_last], base_col=1)

    return row
