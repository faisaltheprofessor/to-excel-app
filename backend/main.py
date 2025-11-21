import io
from fastapi import FastAPI, Request
from fastapi.responses import StreamingResponse

from openpyxl import Workbook

from config import (
    LEFT_HEADERS,
    PERM_HEADERS,
    DATA_START_ROW,
    ROLES_COUNT,
)
from utils import tree_max_depth, autosize_columns
from worksheets import create_sheet, write_rows, add_second_sheet, add_third_sheet
app = FastAPI()


@app.post("/generate-excel")
async def generate_excel(request: Request):
    """
    JSON body:
      {
        "tree": [...],                 # REQUIRED
        "sheets": ["GE","Ablage","Roles"],   # OPTIONAL
        "rolesCount": 10              # OPTIONAL
      }
    """
    data = await request.json()
    tree = data.get("tree")
    if not tree:
        return {"error": "Missing 'tree' in JSON body"}

    sheets = data.get("sheets") or ["GE", "Ablage", "Roles"]
    roles_count = int(data.get("rolesCount") or ROLES_COUNT)

    max_depth = max(0, tree_max_depth(tree))
    last_name_col = max_depth + 1

    if "GE" in sheets:
        wb, ws, cols = create_sheet(
            last_name_col,
            perm1_cols=len(LEFT_HEADERS),
            perm2_cols=len(PERM_HEADERS),
            max_depth=max_depth
        )
        write_rows(
            ws, tree, DATA_START_ROW, level=0, last_stack=[],
            cols=cols, lineage_names=[], lineage_apps=[]
        )
        autosize_columns(ws, min_w=8, max_w=60)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    if "Ablage" in sheets:
        add_second_sheet(wb, tree)

    if "Roles" in sheets:
        add_third_sheet(wb, tree, roles_count)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tree.xlsx"}
    )
